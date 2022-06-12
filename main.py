import tkinter as tk
import xlsxwriter
import os
from openpyxl import load_workbook
from tkinter import *
from math import exp, log, sqrt
from itertools import permutations
import time
from tkinter import messagebox
from random import random, uniform, gauss
import winsound
frequency = 1000  # Set Frequency To 2500 Hertz
duration = 100  # Set Duration To 1000 ms == 1 second
winsound.Beep(frequency, duration)

global msayisi
global ssayisi
global nehgantsira

def kapatma2():
    window0.destroy()

def degergetir():
    global msayisi
    global ssayisi
    sonuc2 = Label(text=str(issayisi.get()) + '   /   ' + str(makinesayisi.get()) + '   /  F  /  Fmax  ',
                   fg='white', bg='purple', width=20, font=196)
    sonuc2.place(x=720, y=100)
    msayisi = int(makinesayisi.get())
    ssayisi = int(issayisi.get())
    if cb2.get() == 1:
        buton_baslat['state'] = NORMAL
        buton_baslat.configure(text='B A Ş L A T', bg='green', fg='white', font=36)
    elif cb2.get() == 0:
        buton_baslat['state'] = DISABLED
        buton_baslat.configure(text='Başlatılamaz!', bg='gray', fg='white')



def sonucac():
    messagebox.showinfo('Bilgilendirme', 'Excel dosyası açılacak. Programın devam edebilmesi için'
                                         ' Excel dosyasını kapatmayı unutmayınız.')
    os.system("sonuclarrr.xlsx")


def GANT(algoritma_adi, sonuc, sayfano, yazmasirasi_2):
    sayac11 = 0
    gliste1 = []
    gliste2 = []
    glistesonuc = []
    isindex = []
    sonliste = []
    toplam = 0
    sayacindex = 0
    bekleme = 0
    zaman = 0
    guncelleme = 0
    sayac9 = 0
    sayac10 = 0
    sayacgant = 3
    yazmasirasi = 0

    while sayac11 < is_sayisi:
        sayfano.write(yazmasirasi, sayac11, sonuc[sayac11], renklendirme2)
        sayac11 += 1
    sayac11 = 0
    yazmasirasi += 1

    for cell in sayfa['2']:
        if cell.value is not None:
            gliste1.append(cell.value)
    gliste1.pop(0)
    while sayacindex < is_sayisi:
        isindex.append(sonuc[sayacindex] - 1)
        toplam = toplam + gliste1[isindex[sayacindex]]
        glistesonuc.append(toplam)
        sayacindex += 1
        # ilk makinenin gant aşaması tamamlandı

    while sayac11 < is_sayisi:
        sayfano.write(yazmasirasi, sayac11, glistesonuc[sayac11])
        sayac11 += 1
    sayac11 = 0
    yazmasirasi += 1

    while sayacgant < makine_sayisi + 2:
        for cell in sayfa[sayacgant]:
            if cell.value is not None:
                gliste2.append(cell.value)
        gliste2.pop(0)
        while sayac9 < is_sayisi:
            while glistesonuc[sayac9] > bekleme:
                bekleme = bekleme + 1
            zaman = bekleme + gliste2[isindex[sayac9]]
            sonliste.append(zaman)
            guncelleme = sonliste[sayac9 - 1] + gliste2[isindex[sayac9]]
            if sonliste[sayac9] < guncelleme:
                sonliste[sayac9] = guncelleme
            if sayac9 == 0:
                sonliste[sayac9] = zaman
            sayac9 += 1
            zaman = 0
            bekleme = 0
        while sayac10 < len(sonliste):
            glistesonuc[sayac10] = sonliste[sayac10]
            sayac10 += 1
        while sayac11 < is_sayisi:
            sayfano.write(yazmasirasi, sayac11, glistesonuc[sayac11])
            sayac11 += 1
        yazmasirasi += 1
        sayac11 = 0
        sonliste.clear()
        # print(glistesonuc)
        sayac9 = 0
        sayac10 = 0
        gliste2.clear()
        sayacgant += 1
    sayacgant = 3
    guncelleme = 0
    sonuc_sira.clear()
    sonuc_sira.append(sonuc)
    sonuc_fmax.clear()
    sonuc_fmax.append(max(glistesonuc))
    sayfa6.write(yazmasirasi_2, 0, algoritma_adi, renklendirme)
    sayfa6.write(yazmasirasi_2, 1, sonuc_fmax[0], renklendirme1)
    sayfa6.write_row(yazmasirasi_2, 2, sonuc_sira[0])


def NEH(yazmasirasi):
    sayac11 = 0
    while sayac11 < len(Neh_gant):
        sayfa5.write(yazmasirasi, sayac11, Neh_gant[sayac11], renklendirme2)
        sayac11 += 1
    sayac11 = 0
    yazmasirasi += 1
    gliste1 = []
    gliste2 = []
    global nehgantsira
    nehgantsira = []
    isindex = []
    sonliste = []
    toplam = 0
    sayacindex = 0
    bekleme = 0
    zaman = 0
    guncelleme = 0
    sayac9 = 0
    sayac10 = 0
    sayacgant = 3
    bitis = makine_sayisi + 1

    # işleri index islemine düzeltme
    for cell in sayfa['2']:
        if cell.value is not None:
            gliste1.append(cell.value)
    gliste1.pop(0)
    while sayacindex < len(Neh_gant):
        isindex.append(Neh_gant[sayacindex] - 1)
        toplam = toplam + gliste1[isindex[sayacindex]]
        nehgantsira.append(toplam)
        sayacindex += 1
        # ilk makinenin gant aşaması tamamlandı
    while sayac11 < len(Neh_gant):
        sayfa5.write(yazmasirasi, sayac11, nehgantsira[sayac11])
        sayac11 += 1
    sayac11 = 0
    yazmasirasi += 1
    # print(glistesonuc)
    while sayacgant < makine_sayisi + 2:
        for cell in sayfa[sayacgant]:
            if cell.value is not None:
                gliste2.append(cell.value)
        gliste2.pop(0)
        while sayac9 < len(Neh_gant):
            while nehgantsira[sayac9] > bekleme:
                bekleme = bekleme + 1
            zaman = bekleme + gliste2[isindex[sayac9]]
            sonliste.append(zaman)
            guncelleme = sonliste[sayac9 - 1] + gliste2[isindex[sayac9]]
            if sonliste[sayac9] < guncelleme:
                sonliste[sayac9] = guncelleme
            if sayac9 == 0:
                sonliste[sayac9] = zaman
            sayac9 += 1
            zaman = 0
            bekleme = 0
        while sayac10 < len(sonliste):
            nehgantsira[sayac10] = sonliste[sayac10]
            sayac10 += 1
        while sayac11 < len(Neh_gant):
            sayfa5.write(yazmasirasi, sayac11, nehgantsira[sayac11])
            sayac11 += 1
        yazmasirasi += 1
        sayac11 = 0
        sonliste.clear()
        sayac9 = 0
        sayac10 = 0
        gliste2.clear()
        sayacgant += 1
    sayacgant = 0
    guncelleme = 0


def SA(yazmasirasi):

    gliste1 = []
    gliste2 = []
    global Sa_fmax
    Sa_fmax = []
    isindex = []
    sonliste = []
    toplam = 0
    sayacindex = 0
    bekleme = 0
    zaman = 0
    guncelleme = 0
    sayac9 = 0
    sayac10 = 0
    sayacgant = 3
    bitis = makine_sayisi + 1

    # işleri index islemine düzeltme
    for cell in sayfa['2']:
        if cell.value is not None:
            gliste1.append(cell.value)
    gliste1.pop(0)
    while sayacindex < len(Simulated_sira):
        isindex.append(Simulated_sira[sayacindex] - 1)
        toplam = toplam + gliste1[isindex[sayacindex]]
        Sa_fmax.append(toplam)
        sayacindex += 1
        # ilk makinenin gant aşaması tamamlandı
    # print(glistesonuc)
    while sayacgant < makine_sayisi + 2:
        for cell in sayfa[sayacgant]:
            if cell.value is not None:
                gliste2.append(cell.value)
        gliste2.pop(0)
        while sayac9 < len(Simulated_sira):
            while Sa_fmax[sayac9] > bekleme:
                bekleme = bekleme + 1
            zaman = bekleme + gliste2[isindex[sayac9]]
            sonliste.append(zaman)
            guncelleme = sonliste[sayac9 - 1] + gliste2[isindex[sayac9]]
            if sonliste[sayac9] < guncelleme:
                sonliste[sayac9] = guncelleme
            if sayac9 == 0:
                sonliste[sayac9] = zaman
            sayac9 += 1
            zaman = 0
            bekleme = 0
        while sayac10 < len(sonliste):
            Sa_fmax[sayac10] = sonliste[sayac10]
            sayac10 += 1
        sayac11 = 0
        sonliste.clear()
        sayac9 = 0
        sayac10 = 0
        gliste2.clear()
        sayacgant += 1
    sayacgant = 0
    guncelleme = 0
    # print(Sa_fmax)


def girdiac():
    messagebox.showinfo('Bilgilendirme', 'Excel dosyası açılacak. Uygun formatta giriş yapınız. '
                                         'Kaydettikten sonra Excel dosyasını kapatmayı unutmayınız.')
    os.system("Girdiler.xlsx")


def calistir():
    messagebox.showinfo("Bilgilendirme", "CDS-PALMER-RAP-GUPTA-NEH Algoritmaları Arkaplanda çözümlenecek.")
    window.destroy()


def Kontrol1():
    if cb.get() == 1:
        buton_devam['state'] = NORMAL
        buton_devam.configure(text='DEVAM', bg='green', fg='white')
    elif cb.get() ==0:
        buton_devam['state'] = DISABLED
        buton_devam.configure(text='Onaylanmadı!', bg='gray', fg='white')


window0 = tk.Tk()
window0.geometry('1200x450+10+10')
window0.configure(bg='black')
window0.resizable(False, False)
window0.title('Akış Tipi Atöyler İçin Sıralama Optimizasyonu')

adim1 = tk.Label(window0, text='1. ADIM: Girdiler Excel dosyasını yandaki', font=24)
adim1.place(x=10, y=10)
adim1_2 = tk.Label(window0, text='formatta doldurup kaydediniz.', font=24)
adim1_2.place(x=10, y=40)
adim1_3 = tk.Label(window0, text='Sayı girişlerini el ile (formül kullanmadan) yapınız', bg='red', fg='white')
adim1_3.place(x=40, y=100)
adim1_4 = tk.Label(window0, text='Sütunlar, İş Sayısını; Satırlar, Makine Sayısını belirtir.', foreground='red',
                   bg='red', fg='white')
adim1_4.place(x=40, y=130)
adim1_4 = tk.Label(window0, text='Tam Sayı tipinde giriş yapınız.', foreground='red', bg='red', fg='white')
adim1_4.place(x=40, y=160)
adim1_5 = tk.Label(window0, text='Sarı kısımlara ve boş kalacak kısımlara herhangi bir işlem yapmayınız.',
                   foreground='red', bg='red', fg='white')
adim1_5.place(x=40, y=190)
adim1_6 = tk.Label(window0, text='Program çalışırken Excel dosyalarını açmayınız.', foreground='red', bg='red',
                   fg='white')
adim1_6.place(x=40, y=220)
adim1_7 = tk.Label(window0, text='Dosyayı kaydedip kapattıktan sonra kutucuğu işaretleyiniz.', font=24)
adim1_7.place(x=10, y=340)

foto = PhotoImage(file="sablon.png")
sablon = Label(window0, image=foto)
sablon.place(x=530, y=10)

buton_girdiler = Button(window0, text='Dosyayı Aç', font=128, bg='green', fg='white', command=girdiac)
buton_girdiler.place(x=30, y=280)
cb = IntVar()
check_girdiler = Checkbutton(window0, text='Verileri Girdim/Onayladım', variable=cb, onvalue=1, offvalue=0,
                             fg='white', bg='blue', font=72, state=NORMAL, selectcolor='red', command=Kontrol1)
check_girdiler.place(x=190, y=280)
buton_devam = Button(window0, text='Onaylanmadı!', height=2, width=20, font=36,
                     bg='gray', fg='red', state=DISABLED, command=kapatma2)
buton_devam.place(x=900, y=380)
window0.mainloop()

kitap = load_workbook('Girdiler.xlsx')
kitap1 = xlsxwriter.Workbook('sonuclarrr.xlsx')
window = tk.Tk()
window.geometry('1000x300+10+10')
window.configure(bg='black')
window.resizable(False, False)
# form.state('normal')  #normal ekran
# form.state('zoomed') #tam ekran
window.title('Akış Tipi Atöyler İçin Sıralama Optimizasyonu')
etiket1 = tk.Label(window, text='Makine (M) Sayısı:', font=96, width=30)
etiket2 = tk.Label(window, text='İş (Job) Sayısı:', font=96, width=30)

makinesayisi = Entry(window, justify='right', fg='blue', bg='yellow', font=128, width=5)
issayisi = Entry(window, justify='right', fg='blue', bg='yellow', font=128, width=5)

adim2 = tk.Label(window, text='2. ADIM: n/m/F/Fmax için Makine Sayısı (m) ve İş Sayısı (n) girip "İŞLE ve'
                              ' GETİR" butonuna tıklayınız.', font=24)
adim2.place(x=10, y=10)
etiket1.place(x=50, y=80)
etiket2.place(x=50, y=120)
sonuc1 = tk.Label(window, text='İşleme Alınan Problem:', font=24)
sonuc1.place(x=500, y=100)
adim3 = tk.Label(window, text='3. ADIM: Algoritmayı Başlatınız.', font=24)
adim3.place(x=10, y=250)
makinesayisi.place(x=420, y=80)
issayisi.place(x=420, y=120)

cb2 = IntVar()
Check_isle = Checkbutton(window, text='İşle ve Getir', variable=cb2, onvalue=1, offvalue=0,
                             fg='white', bg='blue', font=72, state=NORMAL, selectcolor='red', command=degergetir)
Check_isle.place(x=260, y=180)
# buton_isle = Button(window, text='İŞLE VE GETİR', font=128, bg='turquoise', command=degergetir)
# buton_isle.place(x=260, y=180)

buton_baslat = Button(window, text='Başlatılamaz!', font=128, state=DISABLED, bg='gray',
                      fg='blue', width=40, command=calistir)
buton_baslat.place(x=500, y=240)

window.mainloop()

sayfa = kitap.active

sayfa6 = kitap1.add_worksheet('SONUC')
sayfa1 = kitap1.add_worksheet('CDS')
sayfa2 = kitap1.add_worksheet('PALMER')
sayfa3 = kitap1.add_worksheet('RAP')
sayfa4 = kitap1.add_worksheet('GUPTA')
sayfa5 = kitap1.add_worksheet('NEH')


renklendirme = kitap1.add_format()
renklendirme.set_bold(bold=True)
renklendirme.set_bg_color('#FF9999')
renklendirme2 = kitap1.add_format()
renklendirme2.set_bg_color('#CCFFFF')
renklendirme2.set_bold(bold=True)
renklendirme2.set_font_color('#660000')
renklendirme3 = kitap1.add_format()
renklendirme3.set_bg_color('#A0A0A0')
renklendirme3.set_font_size(9)
renklendirme1 = kitap1.add_format()
renklendirme1.set_bg_color('#000066')
renklendirme1.set_font_color('yellow')
renklendirme1.set_bold(bold=True)

sonuc_sira = []
sonuc_fmax = []
yazmasirasi_2 = 1
liste1 = []
liste2 = []
liste3 = []
liste4 = []
liste5 = []
liste6 = []
list1 = []
list2 = []
list3 = []
list4 = []
sonuc = []
sayfano = 0

birlestirveortala = kitap1.add_format({'align': 'center'})
merge_format = kitap1.add_format({'bold': True, 'align': 'center', 'fg_color': '#D7E4BC'})

sayfa6.write(0, 0, 'ALGORİTMA', renklendirme3)
sayfa6.write(0, 1, 'FMAX', renklendirme1)
sayfa6.merge_range('C1:M1', '>>>>>>>>  İ Ş    S I R A L A M A S I >>>>>>>>', merge_format)
is_sayisi = ssayisi
makine_sayisi = msayisi
# CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS CDS
for cell in sayfa['2']:
    if cell.value is not None:
        liste1.append(cell.value)
liste1.pop(0)

for cell in sayfa[makine_sayisi + 1]:
    if cell.value is not None:
        liste2.append(cell.value)
liste2.pop(0)

sayac = 0
islistesi = []
islistesi1 = []
sayac1 = 0
sayac3 = 0
sayac5 = 0
sayac6 = 0
sayac7 = 0
sayac11 = 0
yazmasirasi = 0
islemsayisi = makine_sayisi - 2
islemsayac = 0
secimislistesi = []
secimislistesi1 = []
secilmeyenislistesi = []
secilmeyenislistesi1 = []
secilmeyenlistesi1 = []
secilmeyenlistesi2 = []
secilmeyenlistesi3 = []
secilmeyenlistesi4 = []
baslangic = 2

# sıfırdan bire iş listesi eklemesi yapılır#
while sayac1 < is_sayisi:
    sayac1 += 1
    islistesi.append(sayac1)

while is_sayisi > sayac:
    if min(liste1) < liste2[liste1.index(min(liste1))] or min(liste1) == liste2[liste1.index(min(liste1))]:
        secimislistesi.append(islistesi[liste1.index(min(liste1))])
        del islistesi[liste1.index(min(liste1))]
        del liste2[liste1.index(min(liste1))]
        del liste1[liste1.index(min(liste1))]

    else:
        secilmeyenlistesi1.append(min(liste1))
        secilmeyenlistesi2.append(liste2[liste1.index(min(liste1))])
        secilmeyenislistesi.append(islistesi[liste1.index(min(liste1))])
        del islistesi[liste1.index(min(liste1))]
        del liste2[liste1.index(min(liste1))]
        del liste1[liste1.index(min(liste1))]

    sayac += 1

while len(secilmeyenlistesi2) > 0:
    secimislistesi.append(secilmeyenislistesi[secilmeyenlistesi2.index(max(secilmeyenlistesi2))])
    secilmeyenislistesi.remove(secilmeyenislistesi[secilmeyenlistesi2.index(max(secilmeyenlistesi2))])
    secilmeyenlistesi2.remove(max(secilmeyenlistesi2))

secilmeyenlistesi1.clear()
sonuc_sira.append(secimislistesi)
# print(secimislistesi)
while sayac11 < is_sayisi:
    sayfa1.write(yazmasirasi, sayac11, secimislistesi[sayac11], renklendirme2)
    sayac11 += 1
sayac11 = 0
yazmasirasi += 1
gliste1 = []
gliste2 = []
glistesonuc = []
isindex = []
sonliste = []
toplam = 0
sayacindex = 0
bekleme = 0
zaman = 0
guncelleme = 0
sayac9 = 0
sayac10 = 0
sayacgant = 3
bitis = makine_sayisi + 1

# işleri index islemine düzeltme
for cell in sayfa['2']:
    if cell.value is not None:
        gliste1.append(cell.value)
gliste1.pop(0)
while sayacindex < is_sayisi:
    isindex.append(secimislistesi[sayacindex] - 1)
    toplam = toplam + gliste1[isindex[sayacindex]]
    glistesonuc.append(toplam)
    sayacindex += 1
    # ilk makinenin gant aşaması tamamlandı
while sayac11 < is_sayisi:
    sayfa1.write(yazmasirasi, sayac11, glistesonuc[sayac11])
    sayac11 += 1
sayac11 = 0
yazmasirasi += 1
# print(glistesonuc)
while sayacgant < makine_sayisi + 2:
    for cell in sayfa[sayacgant]:
        if cell.value is not None:
            gliste2.append(cell.value)
    gliste2.pop(0)
    while sayac9 < is_sayisi:
        while glistesonuc[sayac9] > bekleme:
            bekleme = bekleme + 1
        zaman = bekleme + gliste2[isindex[sayac9]]
        sonliste.append(zaman)
        guncelleme = sonliste[sayac9 - 1] + gliste2[isindex[sayac9]]
        if sonliste[sayac9] < guncelleme:
            sonliste[sayac9] = guncelleme
        if sayac9 == 0:
            sonliste[sayac9] = zaman
        sayac9 += 1
        zaman = 0
        bekleme = 0
    while sayac10 < len(sonliste):
        glistesonuc[sayac10] = sonliste[sayac10]
        sayac10 += 1
    while sayac11 < is_sayisi:
        sayfa1.write(yazmasirasi, sayac11, glistesonuc[sayac11])
        sayac11 += 1
    yazmasirasi += 1
    sayac11 = 0
    sonliste.clear()
    # print(glistesonuc)
    sayac9 = 0
    sayac10 = 0
    gliste2.clear()
    sayacgant += 1
sayacgant = 0
guncelleme = 0
sonuc_fmax.append(max(glistesonuc))
sayfa6.write(yazmasirasi_2, 0, 'CDS', renklendirme)
sayfa6.write(yazmasirasi_2, 1, sonuc_fmax[0], renklendirme1)
sayfa6.write_row(yazmasirasi_2, 2, sonuc_sira[0])
yazmasirasi_2 += 1
sonuc_sira.clear()
sonuc_fmax.clear()
# ilk siralama islemleri bitti
for cell in sayfa[baslangic]:
    if cell.value is not None:
        list1.append(cell.value)
list1.pop(0)

for cell in sayfa[makine_sayisi + 1]:
    if cell.value is not None:
        list2.append(cell.value)
list2.pop(0)

while islemsayac < islemsayisi:
    for cell in sayfa[baslangic + 1]:
        if cell.value is not None:
            list3.append(cell.value)
    list3.pop(0)

    for cell in sayfa[bitis - 1]:
        if cell.value is not None:
            list4.append(cell.value)
    list4.pop(0)
    while sayac3 < is_sayisi:
        list1[sayac3] = list1[sayac3] + list3[sayac3]
        list2[sayac3] = list2[sayac3] + list4[sayac3]
        sayac3 += 1
    while sayac7 < is_sayisi:
        liste5.append(list1[sayac7])
        liste6.append(list2[sayac7])
        sayac7 += 1
    sayac7 = 0
    baslangic += 1
    bitis -= 1
    list4.clear()
    list3.clear()
    sayac3 = 0
    islemsayac += 1
    while sayac5 < is_sayisi:
        sayac5 += 1
        islistesi1.append(sayac5)
    sayac5 = 0
    while is_sayisi > sayac6:
        if min(liste5) < liste6[liste5.index(min(liste5))] or min(liste5) == liste6[liste5.index(min(liste5))]:
            secimislistesi1.append(islistesi1[liste5.index(min(liste5))])
            del islistesi1[liste5.index(min(liste5))]
            del liste6[liste5.index(min(liste5))]
            del liste5[liste5.index(min(liste5))]

        else:
            secilmeyenlistesi4.append(liste6[liste5.index(min(liste5))])
            secilmeyenlistesi3.append(liste5[liste5.index(min(liste5))])
            secilmeyenislistesi1.append(islistesi1[liste5.index(min(liste5))])
            del liste6[liste5.index(min(liste5))]
            del islistesi1[liste5.index(min(liste5))]
            del liste5[liste5.index(min(liste5))]

        sayac6 += 1

    sayac6 = 0
    while len(secilmeyenlistesi4) > 0:
        secimislistesi1.append(secilmeyenislistesi1[secilmeyenlistesi4.index(max(secilmeyenlistesi4))])
        secilmeyenislistesi1.remove(secilmeyenislistesi1[secilmeyenlistesi4.index(max(secilmeyenlistesi4))])
        secilmeyenlistesi4.remove(max(secilmeyenlistesi4))
    secilmeyenlistesi3.clear()
    while sayac11 < is_sayisi:
        sayfa1.write(yazmasirasi, sayac11, secimislistesi1[sayac11], renklendirme2)
        sonuc_sira.append(secimislistesi1[sayac11])
        sayac11 += 1
    yazmasirasi += 1
    gliste1 = []
    gliste2 = []
    glistesonuc = []
    isindex = []
    sonliste = []
    toplam = 0
    sayacindex = 0
    bekleme = 0
    zaman = 0
    sayac9 = 0
    sayac10 = 0
    sayacgant = 3
    sayac11 = 0
    for cell in sayfa['2']:
        if cell.value is not None:
            gliste1.append(cell.value)
    gliste1.pop(0)
    while sayacindex < is_sayisi:
        isindex.append(secimislistesi1[sayacindex] - 1)
        toplam = toplam + gliste1[isindex[sayacindex]]
        glistesonuc.append(toplam)
        sayacindex += 1
        # ilk makinenin gant asamasi tamamlandi
    while sayac11 < is_sayisi:
        sayfa1.write(yazmasirasi, sayac11, glistesonuc[sayac11])
        sayac11 += 1
    sayac11 = 0
    yazmasirasi += 1
    # print(glistesonuc)
    while sayacgant < makine_sayisi + 2:
        for cell in sayfa[sayacgant]:
            if cell.value is not None:
                gliste2.append(cell.value)
        gliste2.pop(0)
        while sayac9 < is_sayisi:
            while glistesonuc[sayac9] > bekleme:
                bekleme = bekleme + 1
            zaman = bekleme + gliste2[isindex[sayac9]]
            sonliste.append(zaman)
            guncelleme = sonliste[sayac9 - 1] + gliste2[isindex[sayac9]]
            if sonliste[sayac9] < guncelleme:
                sonliste[sayac9] = guncelleme
            if sayac9 == 0:
                sonliste[sayac9] = zaman
            sayac9 += 1
            zaman = 0
            bekleme = 0
        while sayac10 < len(sonliste):
            glistesonuc[sayac10] = sonliste[sayac10]
            sayac10 += 1
        while sayac11 < is_sayisi:
            sayfa1.write(yazmasirasi, sayac11, glistesonuc[sayac11])
            sayac11 += 1
        yazmasirasi += 1
        sayac11 = 0
        sonliste.clear()
        # print(glistesonuc)
        sayac9 = 0
        sayac10 = 0
        gliste2.clear()
        sayacgant += 1
    sayacgant = 0
    guncelleme = 0
    sonuc_fmax.append(max(glistesonuc))
    secimislistesi1.clear()
    secilmeyenlistesi4.clear()
    secilmeyenlistesi3.clear()
    secilmeyenislistesi1.clear()
    islistesi1.clear()
    sayfa6.write(yazmasirasi_2, 0, 'CDS', renklendirme)
    sayfa6.write(yazmasirasi_2, 1, sonuc_fmax[0], renklendirme1)
    sayfa6.write_row(yazmasirasi_2, 2, sonuc_sira)
    yazmasirasi_2 += 1
    sonuc_sira.clear()
    sonuc_fmax.clear()

# PALMER  PALMER PALMER PALMER PALMER PALMER PALMER PALMER PALMER PALMER PALMER PALMER PALMER PALMER PALMER
yazmasirasi = 0
sayac2 = 0
sayac4 = 0
sayac8 = 0
sayac12 = 2
sayac13 = 0
baslangic = 2
mliste = []
psonuc = []
plistesi = []
lgirdiler = []
palmer = []
psiralama = []
while sayac4 < is_sayisi:
    sayac4 += 1
    plistesi.append(sayac4)
while sayac8 < makine_sayisi:
    sayac8 += 1
    mliste.append(sayac8)
while sayac12 < is_sayisi + 2:
    for column in range(sayac12, sayac12 + 1):
        for row in range(2, makine_sayisi + 2):
            value = sayfa.cell(row, column).value
            lgirdiler.append(value)
    # print(lgirdiler)
    while sayac2 < makine_sayisi:
        palmer.append(- ((makine_sayisi - ((2 * mliste[sayac2]) - 1)) * lgirdiler[sayac2]))
        # print(palmer)
        sayac2 += 1
    lgirdiler.clear()
    sayac12 += 1
    sayac2 = 0
    psonuc.append(sum(palmer))
    palmer.clear()
    # print(psonuc)
while is_sayisi > sayac13:
    cikartilan = psonuc.index(max(psonuc))
    psiralama.append(plistesi[cikartilan])
    plistesi.remove(plistesi[psonuc.index(max(psonuc))])
    psonuc.remove(psonuc[psonuc.index(max(psonuc))])

    # print("--", psiralama)
    sayac13 += 1
# print(plistesi)
# print(psiralama)

GANT('PALMER', psiralama, sayfa2, yazmasirasi_2)

# RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP RAP
yazmasirasi = 0
sayacindex = 0
sayac13 = 0
sayac14 = 0
sayac15 = 2
sayac16 = 0
sayac17 = 0
rlistesi = []
rgirdiler = []
j_liste = []
jliste = []
rap1 = []
rap2 = []
rapsonuc1 = []
rapsonuc2 = []
rapsonuc = []
while sayac13 < is_sayisi:
    sayac13 += 1
    rlistesi.append(sayac13)

while sayac14 < makine_sayisi:
    sayac14 += 1
    j_liste.append(sayac14)
jliste = j_liste[::-1]
while sayac15 < is_sayisi + 2:
    for columnn in range(sayac15, sayac15 + 1):
        for roww in range(2, makine_sayisi + 2):
            value = sayfa.cell(roww, columnn).value
            rgirdiler.append(value)
    # print(lgirdiler)
    while sayac16 < makine_sayisi:
        rap1.append(jliste[sayac16] * rgirdiler[sayac16])
        rap2.append(j_liste[sayac16] * rgirdiler[sayac16])
        sayac16 += 1

    rgirdiler.clear()
    sayac15 += 1
    sayac16 = 0
    rapsonuc1.append(sum(rap1))
    rap1.clear()
    rapsonuc2.append(sum(rap2))
    rap2.clear()
# print(rapsonuc1)
# print(rapsonuc2)

sayac11 = 0
gliste1 = []
gliste2 = []
glistesonuc = []
isindex = []
sonliste = []
toplam = 0
sayacindex = 0
bekleme = 0
zaman = 0
guncelleme = 0
sayac9 = 0
sayac10 = 0
sayacgant = 3

while is_sayisi > sayac17:
    if min(rapsonuc1) < rapsonuc2[rapsonuc1.index(min(rapsonuc1))] \
            or min(rapsonuc1) == rapsonuc2[rapsonuc1.index(min(rapsonuc1))]:
        rapsonuc.append(rlistesi[rapsonuc1.index(min(rapsonuc1))])
        del rlistesi[rapsonuc1.index(min(rapsonuc1))]
        del rapsonuc2[rapsonuc1.index(min(rapsonuc1))]
        del rapsonuc1[rapsonuc1.index(min(rapsonuc1))]

    else:
        secilmeyenlistesi1.append(min(rapsonuc1))
        secilmeyenlistesi2.append(rapsonuc2[rapsonuc1.index(min(rapsonuc1))])
        secilmeyenislistesi.append(rlistesi[rapsonuc1.index(min(rapsonuc1))])
        del rlistesi[rapsonuc1.index(min(rapsonuc1))]
        del rapsonuc2[rapsonuc1.index(min(rapsonuc1))]
        del rapsonuc1[rapsonuc1.index(min(rapsonuc1))]

    sayac17 += 1

while len(secilmeyenlistesi2) > 0:
    rapsonuc.append(secilmeyenislistesi[secilmeyenlistesi2.index(max(secilmeyenlistesi2))])
    secilmeyenislistesi.remove(secilmeyenislistesi[secilmeyenlistesi2.index(max(secilmeyenlistesi2))])
    secilmeyenlistesi2.remove(max(secilmeyenlistesi2))

# print(rapsonuc)
yazmasirasi_2 += 1
GANT('RAP', rapsonuc, sayfa3, yazmasirasi_2)

# GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA GUPTA
yazmasirasi = 0
sayac18 = 0
sayac19 = 2
sayac20 = 0
guptakiyas = 0
guptasayi = 0
mingupta = []
g_is_liste = []
g_girdiler = []
gup_sonuc = []
gup_siralama = []
while sayac18 < is_sayisi:
    sayac18 += 1
    g_is_liste.append(sayac18)

while sayac19 < is_sayisi + 2:
    for columnnn in range(sayac19, sayac19 + 1):
        for rowww in range(2, makine_sayisi + 2):
            value = sayfa.cell(rowww, columnnn).value
            g_girdiler.append(value)

    if g_girdiler[makine_sayisi - 1] > g_girdiler[0]:
        guptakiyas = - 1
    else:
        guptakiyas = 1
    while sayac20 < makine_sayisi - 2:
        mingupta.append(g_girdiler[sayac20] + g_girdiler[sayac20 + 1])
        sayac20 += 1
        guptasayi = guptakiyas / min(mingupta)
    # print(guptasayi)
    gup_siralama.append(guptasayi)
    mingupta.clear()
    g_girdiler.clear()
    sayac19 += 1
    sayac20 = 0
while sayac20 < is_sayisi:
    gup_sonuc.append(g_is_liste[gup_siralama.index(min(gup_siralama))])
    del g_is_liste[gup_siralama.index(min(gup_siralama))]
    del gup_siralama[gup_siralama.index(min(gup_siralama))]
    sayac20 += 1
# print(gup_sonuc)
# print(gup_siralama)
yazmasirasi_2 += 1
GANT('GUPTA', gup_sonuc, sayfa4, yazmasirasi_2)

# NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH NEH
sayac20 = 0
sayac19 = 2
sayac18 = 0
Neh_is = []
Neh_girdiler = []
Neh_top_girdi = []
Neh_sira = []
Neh_gant = []
while sayac18 < is_sayisi:
    sayac18 += 1
    Neh_is.append(sayac18)

while sayac19 < is_sayisi + 2:
    for columnnn in range(sayac19, sayac19 + 1):
        for rowww in range(2, makine_sayisi + 2):
            value = sayfa.cell(rowww, columnnn).value
            Neh_girdiler.append(value)
    Neh_top_girdi.append(sum(Neh_girdiler))
    Neh_girdiler.clear()
    # print(Neh_top_girdi)
    sayac19 += 1

while sayac20 < is_sayisi:
    Neh_sira.append(Neh_is[Neh_top_girdi.index(max(Neh_top_girdi))])
    del Neh_is[Neh_top_girdi.index(max(Neh_top_girdi))]
    del Neh_top_girdi[Neh_top_girdi.index(max(Neh_top_girdi))]
    sayac20 += 1
# print(Neh_sira)

Neh_gant.append(Neh_sira[0])
del Neh_sira[0]
Neh_gant.append(Neh_sira[0])
del Neh_sira[0]

yazmasirasi = 0

NEH(yazmasirasi)
ilkNeh = nehgantsira[1]
Neh_gant.reverse()
NEH(yazmasirasi)
if ilkNeh < nehgantsira[1]:
    Neh_gant.reverse()
    NEH(yazmasirasi)
yazmasirasi_4 = makine_sayisi + 3
sayac21 = 0
sayac22 = 0
exitdongu = -2
kontrolsonuc = []
sayfa5.write(makine_sayisi + 2, 0, 'NEH Sıralama Seçim Sonuçları')
while sayac21 < is_sayisi - 2:

    while sayac22 < sayac21 + 3 and exitdongu == -2:
        Neh_gant.insert(sayac22, Neh_sira[0])
        NEH(yazmasirasi)
        kontrolsonuc.append(nehgantsira[len(Neh_gant) - 1])

        sayac24 = 0

        while sayac24 < len(Neh_gant):
            sayfa5.write(yazmasirasi_4, 0, kontrolsonuc[sayac22], renklendirme)
            sayfa5.write(yazmasirasi_4, sayac24 + 1, Neh_gant[sayac24])
            sayac24 += 1
        yazmasirasi_4 += 1

        del Neh_gant[sayac22]

        if sayac22 == sayac21 + 2:
            # print(sayac22, Neh_gant, kontrolsonuc)
            sayac22 = kontrolsonuc.index(min(kontrolsonuc))
            Neh_gant.insert(sayac22, Neh_sira[0])
            NEH(yazmasirasi)
            exitdongu = -1
        sayac22 += 1
    exitdongu = -2
    # print("bitis", Neh_gant)
    kontrolsonuc.clear()
    del Neh_sira[0]
    sayac22 = 0
    sayac21 += 1
    yazmasirasi_4 += 1
# print(Neh_gant)
yazmasirasi_2 += 1
sayfa6.write(yazmasirasi_2, 0, 'NEH', renklendirme)
sayfa6.write(yazmasirasi_2, 1, nehgantsira[is_sayisi - 1], renklendirme1)
sayac23 = 0
while sayac23 < is_sayisi:
    sayfa6.write(yazmasirasi_2, sayac23 + 2, Neh_gant[sayac23])
    sayac23 += 1
time.sleep(3)
kitap1.close()


# os.system("sonuclarrr.xlsx")

window_3 = tk.Tk()
window_3.geometry('1300x200+150+150')
window_3.configure(bg='black')
window_3.resizable(True, False)
# form.state('normal')  #normal ekran
# form.state('zoomed') #tam ekran
window_3.title(' Tavlama Benzetimi (Simulated Annealing)')
buton_sonuc = Button(window_3, text='Sonuçları Göster', font=128, bg='blue', fg='white', command=sonucac)
buton_sonuc.place(x=150, y=120)
# buton_tavlama = Button(window_3, text='TAVLAMA BENZETİMİ BAŞLAT', font=128, bg='red')


def kapatma():
    messagebox.showinfo("Bilgilendirme", "Tavlama Benzetimi arka planda çözümlenmek üzere başlatılacak.")
    window_3.destroy()


winsound.Beep(frequency, duration)

buton_tavlama = Button(window_3, text='Tavlama Benzetimi Başlat', font=128, fg='White', bg='green', command=kapatma)
buton_tavlama.place(x=850, y=120)

etiket3 = Label(window_3, text='EN İYİ FMAX', fg='red', bg='GRAY', width=12, font=24)
etiket3.place(x=10, y=10)
etiket4 = Label(window_3, text='EN İYİ SIRALAMA', fg='red', bg='GRAY', width=100, font=18, wraplength=1000)
etiket4.place(x=180, y=10)

kitap2 = load_workbook('sonuclarrr.xlsx')
sayfa_sonuc = kitap2.active


sonuc_siralamasi = []
fmax_liste = []

for columnnnn in range(2, 3):
    for rowwww in range(2, makine_sayisi + 5):
        value = sayfa_sonuc.cell(rowwww, columnnnn).value
        if value is not None:
            fmax_liste.append(value)

for cell in sayfa_sonuc[fmax_liste.index(min(fmax_liste)) + 2]:
    if cell.value is not None:
        sonuc_siralamasi.append(cell.value)
del sonuc_siralamasi[0]
iyiFmax = sonuc_siralamasi[0]
del sonuc_siralamasi[0]
# del sonuc_siralamasi[0]

# print(fmax_liste)
# print(iyiFmax)
kitap2.close()


bas_sonuc1 = Label(window_3, text=iyiFmax, fg='black', bg='yellow', width=12, font=32)
bas_sonuc1.place(x=10, y=40)
bas_sonuc = Label(window_3, text=sonuc_siralamasi, fg='green', bg='white', width=100, font=24, wraplength=1000)
bas_sonuc.place(x=180, y=40)

# Tavlama Benzetimi # Tavlama Benzetimi # Tavlama Benzetimi # Tavlama Benzetimi # Tavlama Benzetimi # Tavlama Benzetimi
iterasyon = 0
Simulated_sira = []
# sayac25 = 0
# yazmasirasi_3 = 0
# yazmasirasi_5 = 0
#
# # Tüm olasılık hesabı
# perm = permutations(sonuc_siralamasi)
# sayac26 = 0
#
# for i in list(perm):
#     # print(i)
#     while sayac26 < is_sayisi:
#         Simulated_sira.append(i[sayac26])
#         sayfa7.write(yazmasirasi_5, sayac26, i[sayac26], renklendirme1)
#         sayac26 += 1
#     sayac26 = 0
#     yazmasirasi_5 += 1
#
#     SA(yazmasirasi)
#     print(Simulated_sira)
#     sayfa7.write(yazmasirasi_3, len(sonuc_siralamasi), Sa_fmax[len(Simulated_sira) - 1], renklendirme1)
#     yazmasirasi_3 += 1
#     Simulated_sira.clear()
# # Tüm olasılık hesabı
imposible = 30

# yenilemeli iterasyon
# while iterasyon < 10:
#    iterasyon += 1
#    random.shuffle(Simulated_sira)
#    SA(yazmasirasi)
#
#
# # Başlangıç Çözümü
# olasiliklar = []
# while sayac25 < len(sonuc_siralamasi):
#     Simulated_sira.append(sonuc_siralamasi[sayac25])
#     sayac25 += 1
# perm = permutations(Simulated_sira)
# sayac26 = 0
# for i in list(perm):
#     # print(i)
#     while sayac26 < is_sayisi:
#         olasiliklar.append(i[sayac26])
#         sayac26 += 1
#     sayac26 = 0
# print(olasiliklar)
# while iterasyon < 10:
#    iterasyon += 1
#    random.shuffle(Simulated_sira)
#    SA(yazmasirasi)
# yenilemeli iterasyon sonu

# sonuc_siralamasi   ---> algoritmalar sonucu en iyi sıralama (başlangıç sıralaması)
# SA(yazmasirasi)
# Sa_fmax ----> FMax değeri yani SA(yazmasirasi) fonksiyonun bulduğu değer. En küçükleme yapmak istediğimiz. 53'e yaklaşmalı
# Simulated_sira ------>  fonksiyonda hesaplanacak sıra. aynı zamanda move ile yeri değiştirilecek sira
window_3.mainloop()


window_4 = tk.Tk()
window_4.geometry('1300x250+150+150')
window_4.configure(bg='black')
window_4.resizable(True, False)
# form.state('normal')  #normal ekran
# form.state('zoomed') #tam ekran
window_4.title(' Tavlama Benzetimi (Simulated Annealing)')

etiket3 = Label(window_4, text='MEVCUT FMAX', fg='blue', bg='GRAY', width=12, font=24)
etiket3.place(x=10, y=10)
etiket4 = Label(window_4, text='MEVCUT EN İYİ SIRALAMA:', fg='blue', bg='GRAY', width=100, font=18, wraplength=1000)
etiket4.place(x=180, y=10)
etiket3 = Label(window_4, text='EN İYİ FMAX', fg='white', bg='brown', width=12, font=24)
etiket3.place(x=10, y=110)
etiket4 = Label(window_4, text='TAVLAMA SONUCU BULUNAN EN İYİ SIRALAMA:',
                fg='white', bg='brown', width=100, font=18, wraplength=1000)
etiket4.place(x=180, y=110)

sayac30 = 0
islerin_liste = []
while sayac30 < is_sayisi:               # iş listesi eklemesi yapılır#
    sayac30 += 1
    islerin_liste.append(sayac30)
# print(islerin_liste)
# print(iyiFmax)
fkontrol = iyiFmax
sirakontrol = []
perm = permutations(islerin_liste, 4)
listegol = []
sayac28 = 0
sayac29 = 0
bas_sonuc2 = Label(window_4, text='Tavlama Benzetimi Daha İyi Bir Sıralama Bulamadı', fg='green',
                   bg='white', width=100, font=24, wraplength=1000)
for i in list(perm):
    while sayac28 < 4:
        listegol.append(i[sayac28])
        sayac28 += 1
    farklistesi = list(set(islerin_liste) - set(listegol))
    while sayac29 < len(farklistesi):
        listegol.append(farklistesi[sayac29])
        sayac29 += 1
    # print("sfsf", farklistesi)
    sayac28 = 0
    sayac29 = 0
    Simulated_sira = listegol
    SA(yazmasirasi)
    # print(Simulated_sira ,max(Sa_fmax))
    if fkontrol > max(Sa_fmax):
        fkontrol = max(Sa_fmax)
        sirakontrol = listegol
        # print(listegol)
        bas_sonuc2 = Label(window_4, text=sirakontrol, fg='blue', bg='white', width=100, font=24, wraplength=1000)
        # print('Fmax:', fkontrol, 'Siralama', sirakontrol)
    listegol.clear()
time.sleep(5)
bas_sonuc3 = Label(window_4, text=fkontrol, fg='black', bg='yellow', width=12, font=32)
bas_sonuc3.place(x=10, y=140)
bas_sonuc2.place(x=180, y=140)
bas_sonuc1 = Label(window_4, text=iyiFmax, fg='black', bg='yellow', width=12, font=32)
bas_sonuc1.place(x=10, y=40)
bas_sonuc = Label(window_4, text=sonuc_siralamasi, fg='black', bg='white', width=100, font=24, wraplength=1000)
bas_sonuc.place(x=180, y=40)
##

##

##

## TAVLAMA BAŞLANGIÇ

## Butonsuz tetikleniyor        D U Z E L T DUZELT D U Z E L T DUZELT D U Z E L T

while iterasyon > 20000:
    # Optimize etmek istedigimiz fonksiyon
    GANT('', sonuc, sayfano, yazmasirasi_2)
    # Custom değerler değişiklik aşırı yavaşlatır...
    T = Ti = 10000  # Baslangic sicakligi
    a = 0.95  # adim katsayisi cooling ratio 0.88 - 0.99 arasi seçilir
    max_iter = 2000  # maksimum sogutma sayisi
    n_rep = 20  # hersogutma arasindaki isitma sayisi
    random_starting_point = uniform(3, 9)  # 3 6 arasi
    # rastgele bir değer üret

    # p = random_starting_point # Kullanmadım diğer algoritmalardan iyi sonuçla başlansın diye (rastgele değil)
    cost = GANT(p)  # baslangic cost ifadesi   # diğer algoritmalardan geleni al
    cost = GANT('eniyi')
    swap = [0, is_sayisi - 1]
    swap_sayaci = 0
    if len(swap)> swap_sayaci:
        random.siralamamx2(random(1,9)) # 1 ile 9 arasında komşu değişikliği yap 9dan azsa baştan tekrar say
        # aynisi gelirse bir sağa kay
    else:
        swap.reverse()  # değiştirilen iki fonksiyonu ters çevir HEMEN YAPAR...
    cool_it = 0
    while cool_it < max_iter:  # Sogutma dongusu
        cool_it += 1
        temp_it = 0
        while temp_it < n_rep:  # Isitma dongusu
            temp_it += 1
            # gauss dagilimdan rastgele sayi cek
            current_p = p + gauss(0, T) * log(sqrt(1 + cost))
            # Diger gelebilecek ifadeler
            # p + gauss(0, T)
            # p + gauss(0, T) * cost
            # p + gauss(0, 1) -> T yerine sabit deger de konulabilir.
            # cost'u hesapla Criteria dediğimiz DELTA
            # Kullanılması gereken::: e^(-Delta/T)
            current_cost = current_p
            criteria = current_cost - cost  # Kriteri hesapla
            print("Step: ", cool_it, " p: ", p, " : ", current_cost)
            # Yeni nokta daha iyiyse p'yi guncelle
            if criteria <= 0:
                cost = current_cost
                p = current_p
            else:
                # kotuyse belirli bir olasilikla kotu noktaya gec
                # bu bizi local minima'dan kurtarir.
                # criteria / T pozitif olmak durumunda
                # 1/e^(pozitif) sayi 0'dan sonsuz [1,0) arasi degerler alir
                # random() da [0, 1) arasi uniform dagilimli degerler uretir.
                if random() < exp(-criteria / T):
                    cost = current_cost
                    p = current_p
        T *= a  # sıcaklık güncellemesi. En iyi bu çalışıyor
        # SONLANDIRMA ŞARTLARI
        # iterasyon sayısına ulaşıldıkça sıcaklık sıfıra yaklaşınca
        if T<1:
            break
        # T = exp(-a*cool_it) * Ti # Sogutma ilerledikce guncelleme fonksiyonu farkli olabilir
winsound.Beep(frequency, duration)
window_4.mainloop()


